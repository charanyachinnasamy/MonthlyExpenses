import csv
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import (
    PieChart,
    Series,
    Reference
)
import os
import shutil
from collections import Counter


#Calculating total expenses for each credit card(Discovr/amex.CSV/capitalone/Chase)/Debit card/bank/Paypal
def monthly_expenses_discover(filename,month):
    f = open(filename)
    total_exp = 0
    global expenses
    expenses = {"AMZN":0.0,'GROCERIES':0.0,"Restaurant":0.0,"SAFEWAY":0.0,"FUEL":0.0,"FARMERS MARKET":0.0,"COSTCO":0.0,"PERSONAL":0.0,"CAB":0.0,"RETAIL":0.0,"NETFLIX":0.0,"COMCAST":0.0,"UTILITIES":0.0,'DAYCARE':0.0,"OTHER":0.0}
    for row in csv.reader(f):
        date = row[0]
        month_file = date.split('/')[0]
        if month_file == month:
            if float(row[3]) > 0:
                total_exp = total_exp + float(row[3])
            expenses = expenses_Splitup(row[2],row[3],filename)
    return total_exp,expenses

def monthly_expenses_amex(filename,month):
    f = open(filename)
    total_exp = 0
    global expenses
    expenses = {"AMZN":0.0,'GROCERIES':0.0,"Restaurant":0.0,"SAFEWAY":0.0,"FUEL":0.0,"FARMERS MARKET":0.0,"COSTCO":0.0,"PERSONAL":0.0,"CAB":0.0,"RETAIL":0.0,"NETFLIX":0.0,"COMCAST":0.0,"UTILITIES":0.0,'DAYCARE':0.0,"OTHER":0.0}
    for row in csv.reader(f):
        date = row[0]
        month_file = date.split('/')[0]
        if month_file == month:
            if float(row[2]) > 0:
                total_exp = total_exp + float(row[2])
            expenses = expenses_Splitup(row[1],row[2],filename)
    return total_exp,expenses

def month_expenses_capitalone(filename, month):
    f = open(filename)
    total_exp = 0
    csvreader = csv.reader(f)
    next(csvreader)
    global expenses
    expenses = {"AMZN":0.0,'GROCERIES':0.0,"Restaurant":0.0,"SAFEWAY":0.0,"FUEL":0.0,"FARMERS MARKET":0.0,"COSTCO":0.0,"PERSONAL":0.0,"CAB":0.0,"RETAIL":0.0,"NETFLIX":0.0,"COMCAST":0.0,"UTILITIES":0.0,'DAYCARE':0.0,"OTHER":0.0}
    for row in csv.reader(f):
        if len(row) == 0:
            break
        else:
            date = row[0]
            month_file = date.split('-')[1]
            if month_file == month:
                if len(row[5]):
                    total_exp = total_exp + float(row[5])
                    expenses = expenses_Splitup(row[3],row[5],filename)
    return total_exp,expenses

def monthly_expenses_chase(filename,month):
    f = open(filename)
    total_exp = 0
    global total_credit
    global expenses
    global expenses
    expenses = {"AMZN":0.0,'GROCERIES':0.0,"Restaurant":0.0,"SAFEWAY":0.0,"FUEL":0.0,"FARMERS MARKET":0.0,"COSTCO":0.0,"PERSONAL":0.0,"CAB":0.0,"RETAIL":0.0,"NETFLIX":0.0,"COMCAST":0.0,"UTILITIES":0.0,'DAYCARE':0.0,"OTHER":0.0}
    if 'bank' in filename:
        csvreader = csv.reader(f)
        next(csvreader)
        for row in csv.reader(f):
            date = row[1]
            month_file = date.split('/')[0]
            if month_file == month:
                if not any(value in row[2].upper() for value in ('DISCOVER','CHASE','AMERICAN','CAPITAL ONE')):
                    if float(row[3]) < 0:
                        total_exp = total_exp - float(row[3])
                        expenses = expenses_Splitup(row[2],row[3],filename)
                    else:
                        total_credit = total_credit + float(row[3])
    else:
        for row in csv.reader(f):
            date = row[0]
            month_file = date.split('/')[0]
            if month_file == month:
                if float(row[5]) < 0:
                    total_exp = total_exp - float(row[5])
                    expenses = expenses_Splitup(row[2],row[5],filename)
    return total_exp,total_credit,expenses

def expenses_Splitup(description,cost,filename):
    global expenses
    if 'Chase' in filename:
        if 'AMZN' in description:
            expenses["AMZN"]= expenses["AMZN"] - float(cost)
        elif any(value in description for value in ('GROCERIES','CASH')):
            expenses["GROCERIES"]= expenses["GROCERIES"] - float(cost)
        elif any(value in description for value in ('GRUBHUB','Saravanaa','MADURAI','VISHNUJI','CHAAT','Food' )):
            expenses["Restaurant"]= expenses["Restaurant"] - float(cost)
        elif 'SAFEWAY' in description:
            expenses["SAFEWAY"]=  expenses["SAFEWAY"] - float(cost)
        elif 'OIL' in description:
            expenses["FUEL"]= expenses["FUEL"] - float(cost)
        elif any(value in description for value in ('SQ','FARMS','FARM')):
            expenses["FARMERS MARKET"]= expenses["FARMERS MARKET"] - float(cost)
        elif 'COSTCO' in description:
            expenses["COSTCO"]= expenses["COSTCO"] - float(cost)
        elif any(value in description for value in ('DAISO','ETSY','MACYS','CROCHET','JOANN','AUDIBLE','MICHAEL','GOFNDME','HERRSCHNERS','MARYMAXIM','PAYPAL','Clips')):
            expenses["PERSONAL"]= expenses["PERSONAL"] - float(cost)
        elif any(value in description for value in ('UBER','LYFT')):
            expenses["CAB"]= expenses["CAB"] - float(cost)
        elif any(value in description for value in ('TARGET','WALMART')):
            expenses["RETAIL"]= expenses["RETAIL"] - float(cost)
        elif 'NETFLIX.COM' in description:
            expenses["NETFLIX"]= expenses["NETFLIX"] - float(cost)
        elif 'COMCAST' in description:
            expenses["COMCAST"]= expenses["COMCAST"] - float(cost)
        elif 'PGANDE' in description:
            expenses["UTILITIES"]= expenses["UTILITIES"] - float(cost)
        elif 'CHECK' in description:
            expenses["DAYCARE"]= expenses["DAYCARE"] - float(cost)
        else:
            if float(cost) < 0:
                expenses["OTHER"]= expenses["OTHER"] - float(cost)
        return expenses
    else:
        if 'AMZN' in description:
            expenses["AMZN"]= expenses["AMZN"] + float(cost)
        elif any(value in description for value in ('GROCERIES','CASH')):
            expenses["GROCERIES"]= expenses["GROCERIES"] + float(cost)
        elif any(value in description for value in ('GRUBHUB','Saravanaa','MADURAI','VISHNUJI','CHAAT','Food' )):
            expenses["Restaurant"]= expenses["Restaurant"] + float(cost)
        elif 'SAFEWAY' in description:
            expenses["SAFEWAY"]=  expenses["SAFEWAY"] + float(cost)
        elif 'OIL' in description:
            expenses["FUEL"]= expenses["FUEL"] + float(cost)
        elif any(value in description for value in ('SQ','FARMS','FARM')):
            expenses["FARMERS MARKET"]= expenses["FARMERS MARKET"] + float(cost)
        elif 'COSTCO' in description:
            expenses["COSTCO"]= expenses["COSTCO"] + float(cost)
        elif any(value in description for value in ('DAISO','ETSY','MACYS','CROCHET','JOANN','AUDIBLE','MICHAEL','GOFNDME','HERRSCHNERS','MARYMAXIM','PAYPAL','Clips')):
            expenses["PERSONAL"]= expenses["PERSONAL"]+ float(cost)
        elif any(value in description for value in ('UBER','LYFT')):
            expenses["CAB"]= expenses["CAB"] + float(cost)
        elif any(value in description for value in ('TARGET','WALMART')):
            expenses["RETAIL"]= expenses["RETAIL"] + float(cost)
        elif 'NETFLIX.COM' in description:
            expenses["NETFLIX"]= expenses["NETFLIX"] + float(cost)
        elif 'COMCAST' in description:
            expenses["COMCAST"]= expenses["COMCAST"] + float(cost)
        elif 'PGANDE' in description:
            expenses["UTILITIES"]= expenses["UTILITIES"] + float(cost)
        elif 'CHECK' in description:
            expenses["DAYCARE"]= expenses["DAYCARE"] + float(cost)
        else:
            if float(cost) >0:
                expenses["OTHER"]= expenses["OTHER"] + float(cost)
        return expenses

#Appending to the existing csv file
def create_csv(filename,result,file_to_write):
    df = pd.DataFrame({'Card_Name':[filename],'Card_Expenses':[result]})
    with pd.ExcelWriter(file_to_write,engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name='TotalExpense',startrow=1, header=False)
        workbook = writer.book
        worksheet = writer.sheets["TotalExpense"]
        header_format = workbook.add_format(
        {
            "bold": True,
            "text_wrap": True,
            "valign": "top",
            "fg_color": "#D7E4BC",
            "border": 1,
        }
        )
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num + 1, value, header_format)

#Creating csv file to write the expenses on file
def write_to_csv(filename,result,file_to_write):
    newdata = [" ",filename,result]
    wb = load_workbook(file_to_write)
    ws= wb.worksheets[0]
    ws.append(newdata)
    wb.save(filename=file_to_write)
    wb.close()

def create_csv_expensesSplitup(expenses,file_to_write,total_credit):
    df = {'Individual_Expenses':['Rent'],'Card_Expenses':['3295.00']}
    for key , value in expenses.items():
        if key == 'OTHER':
            if value > 3295.00:
                value = value - 3295.00
            else:
                value = 3295.00 - value
        df['Individual_Expenses'].append(key)
        df['Card_Expenses'].append(value)
    df_new = pd.DataFrame(df)
    with pd.ExcelWriter(file_to_write,engine="openpyxl", mode='a') as writer1:
        df_new.to_excel(writer1, sheet_name='exSplitup',index=False)
    wb = openpyxl.load_workbook(file_to_write)
    sheets = wb.sheetnames
    ws = wb[sheets[1]]
    pie = PieChart()
    labels = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=15)
    data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=15)
    series = Series(data, title_from_data=True)
    pie.append(series)
    pie.set_categories(labels)
    ws.add_chart(pie, "D1")
    wb.save(file_to_write)
    # ws = wb.active
    # ws = wb.get_active_sheet()
    # writer = pd.ExcelWriter(file_to_write)
    # sheet_name = 'pieChart'
    # df_new.to_excel(writer, sheet_name=sheet_name)
    # max_row = len(df)
    # cell_range = xl_range(1, 1, max_row, 1)
    # workbook = writer.book
    # worksheet = writer.sheets[sheet_name]
    # chart = wb.add_chart({'type': 'pie'})
    # chart.add_series({
    #     'categories': '=Sheet1!A2:A17',
    #     'values':     '=Sheet1!B2:B17'
    #  })
    # worksheet.insert_chart('E0', chart)
    # wb.close()
def create_piechart(file_to_write):
    df=pd.read_excel(file_to_write,sheet_name='sheet2')
    wb = openpyxl.load_workbook(file_to_write)
    worksheet = wb.get_sheet_by_name()


#Main program to get the input
if __name__ == '__main__':
    month = input("Enter the month")
    path = os.getcwd()
    file_path = path+"/Input files/"
    filenames = os.listdir(file_path)
    file_to_write = path + "/Output files/"+"expenses"+"_"+month+".xlsx"
    file_to_archive=path + "/Archive/"+"expenses"+"_"+month+".xlsx"
    if os.path.isfile(file_to_write):
        shutil.move(file_to_write,file_to_archive)
    total_exp_month=0
    total_credit = 0
    final_expenses = expenses = {"AMZN":0.0,'GROCERIES':0.0,"Restaurant":0.0,"SAFEWAY":0.0,"FUEL":0.0,"FARMERS MARKET":0.0,"COSTCO":0.0,"PERSONAL":0.0,"CAB":0.0,"RETAIL":0.0,"NETFLIX":0.0,"COMCAST":0.0,"UTILITIES":0.0,'DAYCARE':0.0,"OTHER":0.0}
    for filename in filenames:
        file_to_open = file_path+filename
        if "Discover" in filename:
            result,expenses = monthly_expenses_discover(file_to_open,month)
            if os.path.isfile(file_to_write):
                write_to_csv(filename.split('.')[0],result,file_to_write)
            else:
                create_csv(filename.split('.')[0],result,file_to_write)
            total_exp_month = total_exp_month + result
            final_expenses = Counter(final_expenses) + Counter(expenses)
        elif 'amex' in filename:
            result,expenses = monthly_expenses_amex(file_to_open,month)
            if os.path.isfile(file_to_write):
                write_to_csv(filename.split('.')[0],result,file_to_write)
            else:
                create_csv(filename.split('.')[0],result,file_to_write)
            total_exp_month = total_exp_month + result
            final_expenses = Counter(final_expenses) + Counter(expenses)
        elif 'Capitalone' in filename:
            result, expenses = month_expenses_capitalone(file_to_open,month)
            if os.path.isfile(file_to_write):
                write_to_csv(filename.split('.')[0],result,file_to_write)
            else:
                create_csv(filename.split('.')[0],result,file_to_write)
            total_exp_month = total_exp_month + result
            final_expenses = Counter(final_expenses) + Counter(expenses)
        elif 'Chase' in filename:
            result_debit,result_credit,expenses = monthly_expenses_chase(file_to_open,month)
            total_credit = result_credit
            if os.path.isfile(file_to_write):
                write_to_csv(filename.split('.')[0],result_debit,file_to_write)
            else:
                create_csv(filename.split('.')[0],result_debit,file_to_write)
            total_exp_month = total_exp_month + result_debit
            final_expenses = Counter(final_expenses) + Counter(expenses)
    write_to_csv("Total Expenses for the month",total_exp_month,file_to_write)
    write_to_csv("Total Income for the month",total_credit,file_to_write)
    create_csv_expensesSplitup(final_expenses,file_to_write,total_credit)
    #create_piechart(file_to_write)








